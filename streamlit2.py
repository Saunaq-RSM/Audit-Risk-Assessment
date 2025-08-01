# streamlit_app.py
"""
Streamlit frontend for the RAG-based retrieval script.
Allows users to drag-and-drop public/client PDFs and the Input GPT.xlsx,
and then runs the full retrieval and LLM-answer pipeline, displaying results
and offering the output .xlsx for download.
"""
import streamlit as st
import pandas as pd
import tempfile
import os
import io
import faiss
import numpy as np
import pdfplumber
import re
import requests
import tiktoken
from typing import Dict, List, Tuple
import time
import math
import openpyxl
from io import BytesIO

# -- Configuration from secrets --
LLM_ENDPOINT = st.secrets["AZURE_API_ENDPOINT"]
LLM_API_KEY = st.secrets["AZURE_API_KEY"]
EMBED_ENDPOINT = st.secrets["EMBEDDING_ENDPOINT"]
EMBED_API_KEY  = LLM_API_KEY

ENC = tiktoken.get_encoding("cl100k_base")

# <-- Add initial_sidebar_state to collapse sidebar on start -->
st.set_page_config(
    page_title="RAG Audit Assistant",
    layout="wide",
    initial_sidebar_state="collapsed"
)

st.title("RAG Audit Assistant 🌐🔍")
st.write("Upload your public & client PDFs and the Input GPT.xlsx. Click **Run** to retrieve relevant chunks & generate answers.")

with st.sidebar:
    st.header("Settings")
    chunk_size    = st.number_input("Chunk size (words)",   min_value=50,   max_value=1000, value=200, step=50)
    chunk_overlap = st.number_input("Chunk overlap (words)",min_value=0,    max_value=chunk_size-1, value=50, step=10)
    top_k         = st.number_input("Chunks per query (k)", min_value=1,    max_value=20,   value=5,   step=1)

public_files = st.file_uploader("Public PDFs", type="pdf", accept_multiple_files=True, key="pubs")
client_files = st.file_uploader("Client PDFs", type="pdf", accept_multiple_files=True, key="clients")
excel_file   = st.file_uploader("Input GPT Excel", type="xlsx", key="excel")
if excel_file and 'excel_bytes' not in st.session_state:
    st.session_state.excel_bytes = excel_file.read()

sheet1_name = sheet2_name = None
if 'excel_bytes' in st.session_state:
    wb = openpyxl.load_workbook(BytesIO(st.session_state.excel_bytes), read_only=True)
    sheet_names = wb.sheetnames
    wb.close()
    st.write("**Step 1:** Select sheets from the workbook:")
    sheet1_name = st.selectbox("Sheet for Code 1300 questions", sheet_names, key="sheet1")
    sheet2_name = st.selectbox("Sheet for Overview of Risks", sheet_names, key="sheet2")
    if sheet1_name == sheet2_name:
        st.error("Please select two different sheets.")
        sheet2_name = None

run_button = st.button("Run Retrieval & Generate Answers")

# --- Helper functions (unchanged) ---
def clean_text(raw: str) -> str:
    text = re.sub(r'[\x00-\x1F\x7F]+', ' ', raw)
    text = re.sub(r'\s+', ' ', text).strip()
    return text

def load_main_body(file) -> str:
    pieces = []
    with pdfplumber.open(file) as pdf:
        for page in pdf.pages:
            x0,y0,x1,y1 = page.bbox
            crop = (x0+50, y0+50, x1-50, y1-50)
            txt = page.within_bbox(crop).extract_text() or ""
            pieces.append(txt)
    full = " ".join(pieces)
    return clean_text(full)

def chunk_text_words(text: str, chunk_size: int, chunk_overlap: int) -> List[str]:
    words = text.split()
    stride = chunk_size - chunk_overlap
    chunks = []
    for i in range(0, len(words), stride):
        seg = words[i:i+chunk_size]
        if seg:
            chunks.append(" ".join(seg))
    return chunks

def embed_texts(texts: List[str]) -> np.ndarray:
    headers = {"Content-Type": "application/json", "api-key": EMBED_API_KEY}
    r = requests.post(EMBED_ENDPOINT, headers=headers, json={"input": texts})
    r.raise_for_status()
    embs = [d.get("embedding") for d in r.json().get("data", [])]
    arr = np.array(embs, dtype=np.float32)
    faiss.normalize_L2(arr)
    return arr

def index_documents(docs: Dict[str,str], chunk_size: int, chunk_overlap: int):
    all_chunks, meta = [], []
    for name, text in docs.items():
        chunks = chunk_text_words(text, chunk_size, chunk_overlap)
        for idx,c in enumerate(chunks):
            all_chunks.append(c)
            meta.append((name, idx))
    if not all_chunks:
        return None, [], []
    embeddings = embed_texts(all_chunks)
    index = faiss.IndexFlatIP(embeddings.shape[1])
    index.add(embeddings)
    return index, meta, all_chunks

def retrieve_chunks(query: str, index, meta, chunks, k: int):
    if index is None:
        return []
    q_emb = embed_texts([query])[0]
    D,I = index.search(np.array([q_emb]), k)
    return [(meta[i][0], chunks[i]) for i in I[0]]

def get_llm_response(prompt: str, context: str) -> str:
    headers = {"Content-Type":"application/json","api-key":LLM_API_KEY}
    messages=[{"role":"system","content":(
                "Please respond in natural, flowing English paragraphs. Do not use any markdown syntax. "
                "You are a world-class corporate analysis assistant for an expert audit team. "
                "Use the context to answer due diligence questions.\n" + context
            )},
              {"role":"user","content":prompt}]
    r = requests.post(LLM_ENDPOINT, headers=headers, json={"messages":messages})
    r.raise_for_status()
    time.sleep(3)
    return r.json()["choices"][0]["message"]["content"].strip()

# --- Run pipeline ---
if run_button:
    if not excel_file:
        st.error("Please upload the Input GPT.xlsx file.")
    elif not sheet1_name or not sheet2_name:
        st.error("Please select two distinct sheets before running.")
    else:
        # read bytes again (necessary because we read once above)
        original_bytes = st.session_state.excel_bytes
        df1 = pd.read_excel(BytesIO(original_bytes), sheet_name=sheet1_name)
        df2 = pd.read_excel(BytesIO(original_bytes), sheet_name=sheet2_name)


        # Proceed with pipeline using df1 and df2...
        # rest of code unchanged, but replace hardcoded sheet references
        # e.g., use df1, df2 and write back using ws1 = wb[sheet1_name], ws2 = wb[sheet2_name]

        st.success("Sheets loaded: {} and {}".format(sheet1_name, sheet2_name))

        # Prepare the second‐sheet columns
        cols = ['Fraud Risk Factor?','Internal Controls','Likelihood','Likelihood Explanation',
                'Material Quantitative Impact?','Impact Explanation','Conclusion','SR?']
        prompts = {
            'Fraud Risk Factor?':       "The Fraud Risk Factor of the above risk type. Answer ONLY with Yes or No.",
            'Internal Controls':        "What are the internal controls within the company against this type of risk. Answer with a maximum of 10 words.",
            'Likelihood':               "Based on public and previous sources. What is the likelihood of this type of risk occurring. Answer ONLY with High or Low.",
            'Likelihood Explanation':   "Based on public and previous sources. Only include justification, max 15 words.",
            'Material Quantitative Impact?': "Based on public and previous sources. Estimated impact of this type of risk. Answer ONLY with High or Low.",
            'Impact Explanation':       "Based on public and previous sources. Only include justification, max 15 words.",
            'Conclusion':               "Explain if there is significant risk or if further discussion is needed. Max 10-15 words.",
            'SR?':                      "Answer only with 'SR' or 'No SR'."
        }
        # Ensure those columns exist
        for c in cols:
            if c not in df2.columns:
                df2[c] = ""

        # Calculate total number of LLM calls (sheet1 rows + sheet2 rows * number of prompts)
        total = len(df1) + len(df2) * len(cols)
        pb = st.progress(0.0)
        pt = st.empty()
        processed = 0

        # Prepare and index all docs once
        pub_docs = {f.name: load_main_body(f) for f in public_files}
        cli_docs = {f.name: load_main_body(f) for f in client_files}
        pub_idx, pub_meta, pub_chunks = index_documents(pub_docs, chunk_size, chunk_overlap)
        cli_idx, cli_meta, cli_chunks = index_documents(cli_docs, chunk_size, chunk_overlap)

        # Phase 1: sheet 1
        df1[['Generated answer','Public chunks','Client chunks']] = ""
        for i, row in df1.iterrows():
            if math.isnan(row["#"]):
                continue
            prompt = row['Question']
            example = row['Best practice answer']

            hits_pub = retrieve_chunks(prompt, pub_idx, pub_meta, pub_chunks, top_k)
            hits_cli = retrieve_chunks(prompt, cli_idx, cli_meta, cli_chunks, top_k)
            ctx = "PUBLIC CONTEXT:\n" + "\n".join(f"[{s}] {t}" for s,t in hits_pub)
            ctx += "\n\nCLIENT CONTEXT:\n" + "\n".join(f"[{s}] {t}" for s,t in hits_cli)
            ctx += "\n\nFORMAT EXAMPLE:\n Q: {prompt}\n A: {example}"
            ctx += "\nDO NOT RESPOND WITH MARKDOWN"

            ans = get_llm_response(prompt, ctx)
            df1.at[i, 'Generated answer'] = ans

            processed += 1
            pb.progress(processed / total)
            pt.text(f"Answered {processed} of {total} questions")

        # Phase 2: sheet 2
        for idx, row in df2.iterrows():
            irf = row['Inherent Risk Factor']
            for c in cols:
                base_prompt = f"For the risk type '{irf}', answer: {prompts[c]}"
                hits_pub = retrieve_chunks(base_prompt, pub_idx, pub_meta, pub_chunks, top_k)
                hits_cli = retrieve_chunks(base_prompt, cli_idx, cli_meta, cli_chunks, top_k)
                ctx = "PUBLIC CONTEXT:\n" + "\n".join(f"[{s}] {t}" for s,t in hits_pub)
                ctx += "\n\nCLIENT CONTEXT:\n" + "\n".join(f"[{s}] {t}" for s,t in hits_cli)
                ctx += "\nDO NOT RESPOND WITH MARKDOWN"
                tr_ans = get_llm_response(base_prompt, ctx)
                df2.at[idx, c] = tr_ans
                processed += 1
                pb.progress(processed / total)
                pt.text(f"Answered {processed} of {total} questions")

        # Write both sheets back into the workbook
        wb = openpyxl.load_workbook(BytesIO(original_bytes))

        # Sheet1
        ws1 = wb[wb.sheetnames[0]]
        for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, max_col=len(df1.columns)):
            for cell in row:
                cell.value = None
        for r, row_vals in enumerate(df1.values, start=2):
            for c, val in enumerate(row_vals, start=1):
                ws1.cell(row=r, column=c).value = val

        # Sheet2 (assuming its tab is called 'Overview of risks')
        ws2 = wb['Overview of risks']
        for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, max_col=len(df2.columns)):
            for cell in row:
                cell.value = None
        for r, row_vals in enumerate(df2.values, start=2):
            for c, val in enumerate(row_vals, start=1):
                ws2.cell(row=r, column=c).value = val

        # Output
        out = BytesIO()
        wb.save(out)
        out.seek(0)
        st.success("All done!")
        st.download_button(
            "Download workbook with answers",
            data=out,
            file_name="questions_with_translations.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
